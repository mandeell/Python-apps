import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import json

# Load your JSON from file or paste directly
with open("training_plan.json", "r") as f:
    data = json.load(f)

# Extract headers and rows
headers = data["headers"]
rows = data["rows"]

# Convert list of dicts to DataFrame using only specified headers
df = pd.DataFrame([{header: row.get(header, "") for header in headers} for row in rows])

# Save to Excel with openpyxl engine
excel_path = "Project_Management_Training_Plan.xlsx"
df.to_excel(excel_path, index=False, engine='openpyxl')

# Load workbook for formatting
wb = load_workbook(excel_path)
ws = wb.active

# Apply wrap text and auto column width
for col_idx, col in enumerate(ws.columns, 1):
    max_length = 0
    col_letter = get_column_letter(col_idx)

    for cell in col:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Determine max length for column width
        if cell.value:
            cell_length = len(str(cell.value))
            if cell_length > max_length:
                max_length = cell_length

    # Set column width based on content
    ws.column_dimensions[col_letter].width = min(max_length + 5, 70)  # limit max width

# Save the formatted Excel file
wb.save(excel_path)

print(f"Excel file saved and formatted: {excel_path}")
